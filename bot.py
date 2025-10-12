import os
import logging
import json
import pandas as pd
from datetime import datetime
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler
import yadisk
import openpyxl
from openpyxl import Workbook

# ===== КОНФИГУРАЦИЯ =====
YANDEX_OAUTH_TOKEN = "y0__xD5tej_ARjX1jog-e7myhRxwrvJbwXpXE88VM7cCF1IsDNp8Q"
YANDEX_CLIENT_ID = "bf718202b2c144278d6c98d1634965b3"
YANDEX_CLIENT_SECRET = "89fc85cdcbd04a86bbf18a18fc2472b2"
TELEGRAM_BOT_TOKEN = "8292656109:AAEnkHukfVcHGW7v8iM0yBmxTX_CyMu7vHg"

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Состояния для ConversationHandler
(
    WAITING_FOR_FULLNAME, 
    WAITING_FOR_AGREEMENT_DATE, 
    WAITING_FOR_CLIENT_INFO,
    WAITING_FOR_CASE_TYPE,
    WAITING_FOR_QUERY_DATE,
    WAITING_FOR_RECIPIENT
) = range(6)

# ===== КЛАСС ДЛЯ РАБОТЫ С БАЗОЙ ДАННЫХ ПОЛЬЗОВАТЕЛЕЙ =====
class UserDatabase:
    def __init__(self, yandex_disk, filename='user_database.json'):
        self.y = yandex_disk
        self.filename = 'KAPObot/' + filename
        self.ensure_directory_exists()
        self.load_data()
    
    def ensure_directory_exists(self):
        """Создает папку KAPObot если она не существует"""
        if not self.y.exists('KAPObot'):
            self.y.mkdir('KAPObot')
    
    def load_data(self):
        try:
            # Пытаемся скачать файл с Яндекс.Диска
            self.y.download(self.filename, 'temp_user_db.json')
            with open('temp_user_db.json', 'r', encoding='utf-8') as f:
                self.data = json.load(f)
            os.remove('temp_user_db.json')
        except Exception as e:
            logger.warning(f"Не удалось загрузить базу пользователей: {e}. Создаем новую.")
            self.data = {'users': {}, 'counters': {'agreement': 0, 'query': 0}}
            self.save_data()
    
    def save_data(self):
        try:
            with open('temp_user_db.json', 'w', encoding='utf-8') as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
            
            # Загружаем на Яндекс.Диск
            with open('temp_user_db.json', 'rb') as f:
                self.y.upload(f, self.filename, overwrite=True)
            
            os.remove('temp_user_db.json')
            logger.info("База пользователей сохранена на Яндекс.Диск")
        except Exception as e:
            logger.error(f"Ошибка при сохранении базы пользователей: {e}")
    
    def is_first_user(self):
        return len(self.data['users']) == 0
    
    def user_exists(self, user_id):
        return str(user_id) in self.data['users']
    
    def add_user(self, user_id, full_name):
        user_id_str = str(user_id)
        role = 'admin' if self.is_first_user() else 'user'
        
        self.data['users'][user_id_str] = {
            'full_name': full_name,
            'role': role,
            'is_blocked': False
        }
        self.save_data()
        return role
    
    def get_user(self, user_id):
        return self.data['users'].get(str(user_id))
    
    def is_admin(self, user_id):
        user = self.get_user(user_id)
        return user and user['role'] == 'admin' and not user['is_blocked']
    
    def get_all_users(self):
        return self.data['users']
    
    def get_next_counter(self, doc_type):
        if doc_type not in self.data['counters']:
            self.data['counters'][doc_type] = 0
        self.data['counters'][doc_type] += 1
        self.save_data()
        return self.data['counters'][doc_type]

# ===== КЛАСС ДЛЯ ЛОГИРОВАНИЯ ДЕЙСТВИЙ =====
class ActionLogger:
    def __init__(self, yandex_disk):
        self.y = yandex_disk
        self.log_filename = 'KAPObot/SoglasheniyaZaprosy/backup/actions.log'
        self.ensure_directory_structure()
    
    def ensure_directory_structure(self):
        """Создает структуру папок на Яндекс.Диске"""
        folders = ['KAPObot', 'KAPObot/SoglasheniyaZaprosy', 'KAPObot/SoglasheniyaZaprosy/backup']
        
        for folder in folders:
            if not self.y.exists(folder):
                self.y.mkdir(folder)
                logger.info(f"Создана папка: {folder}")
    
    def log_action(self, user_id, user_name, action):
        """Логирует действие пользователя"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"{timestamp} - {user_name} (ID: {user_id}) - {action}\n"
        
        try:
            # Скачиваем текущий лог-файл если существует
            temp_log = 'temp_actions.log'
            if self.y.exists(self.log_filename):
                self.y.download(self.log_filename, temp_log)
                mode = 'a'
            else:
                mode = 'w'
            
            # Добавляем запись в лог
            with open(temp_log, mode, encoding='utf-8') as f:
                f.write(log_entry)
            
            # Загружаем обратно на Яндекс.Диск
            with open(temp_log, 'rb') as f:
                self.y.upload(f, self.log_filename, overwrite=True)
            
            os.remove(temp_log)
            logger.info(f"Действие записано в лог: {action}")
            
        except Exception as e:
            logger.error(f"Ошибка при записи в лог: {e}")

# ===== КЛАСС ДЛЯ РАБОТЫ С EXCEL ЧЕРЕЗ YANDEX DISK =====
class ExcelManager:
    def __init__(self, yandex_disk):
        self.y = yandex_disk
        self.filename = "KAPObot/journal_registration.xlsx"  # Файл теперь в корне KAPObot
        self.ensure_directory_structure()
        self.ensure_file_exists()
    
    def ensure_directory_structure(self):
        """Создает структуру папок на Яндекс.Диске"""
        folders = ['KAPObot', 'KAPObot/SoglasheniyaZaprosy', 'KAPObot/SoglasheniyaZaprosy/backup']
        
        for folder in folders:
            if not self.y.exists(folder):
                self.y.mkdir(folder)
                logger.info(f"Создана папка: {folder}")
    
    def ensure_file_exists(self):
        if not self.y.exists(self.filename):
            self._create_new_journal()
    
    def _create_new_journal(self):
        wb = Workbook()
        
        # Создаем листы для разных типов документов
        sheets = {
            'Журнал регистрации Соглашений': [
                ['№ п/п', 'Рег. №', 'Дата заключения соглашения', 'Дата регистрации соглашения',
                 'Ф.И.О. адвоката(-ов) заключившего соглашение', 'Сведения о доверителе',
                 'Характер поручения']  # Убрана графа ФИО доверителя, добавлена графа Характер поручения
            ],
            'Журнал регистрации Адвокатских запросов': [
                ['№ п/п', 'Рег. № адвокатского запроса', 'Дата регистрации адвокатского запроса',
                 'Наименование адресата', 'Фамилия, имя, отчество адвоката, направившего запрос']
            ]
        }
        
        # Удаляем лист по умолчанию и создаем нужные
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        for sheet_name, headers in sheets.items():
            ws = wb.create_sheet(sheet_name)
            ws.append(headers[0])
        
        # Сохраняем локально и загружаем на Яндекс.Диск
        wb.save('temp.xlsx')
        with open('temp.xlsx', 'rb') as f:
            self.y.upload(f, self.filename)
        os.remove('temp.xlsx')
        logger.info("Создан новый журнал на Яндекс.Диске")
    
    def _get_next_row_number(self, worksheet_name):
        try:
            self.y.download(self.filename, 'temp.xlsx')
            wb = openpyxl.load_workbook('temp.xlsx')
            ws = wb[worksheet_name]
            next_row = ws.max_row
            wb.save('temp.xlsx')
            return next_row
        except Exception as e:
            logger.error(f"Ошибка при получении номера строки: {e}")
            return 2
    
    def add_record(self, worksheet_name, data):
        try:
            self.y.download(self.filename, 'temp.xlsx')
            wb = openpyxl.load_workbook('temp.xlsx')
            ws = wb[worksheet_name]
            
            next_row = self._get_next_row_number(worksheet_name)
            data_with_number = [next_row] + data  # № п/п начинается с 1 (next_row уже содержит правильный номер)
            ws.append(data_with_number)
            
            wb.save('temp.xlsx')
            with open('temp.xlsx', 'rb') as f:
                self.y.upload(f, self.filename, overwrite=True)
            
            os.remove('temp.xlsx')
            logger.info(f"Запись добавлена в {worksheet_name}")
            return True
        except Exception as e:
            logger.error(f"Ошибка при добавлении записи: {e}")
            return False

# ===== ИНИЦИАЛИЗАЦИЯ =====
yandex_disk = yadisk.YaDisk(YANDEX_CLIENT_ID, YANDEX_CLIENT_SECRET, YANDEX_OAUTH_TOKEN)
user_db = UserDatabase(yandex_disk)
action_logger = ActionLogger(yandex_disk)
excel_manager = ExcelManager(yandex_disk)

# ===== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ =====
def generate_reg_number(doc_type):
    now = datetime.now()
    counter = user_db.get_next_counter(doc_type)
    prefix = {
        'agreement': 'С',
        'query': 'З'
    }.get(doc_type, '')
    
    return f"{counter}/{now.month:02d}/{now.year % 100:02d}-{prefix}"

def get_main_menu_keyboard():
    keyboard = [
        ["📄 Соглашение", "📨 Адвокатский запрос"]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_today_button_keyboard():
    keyboard = [["📅 Сегодня"], ["🔙 Назад"]]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_back_keyboard():
    keyboard = [["🔙 Назад"]]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

# ===== ОСНОВНЫЕ ОБРАБОТЧИКИ =====
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    if user_db.user_exists(user_id):
        user = user_db.get_user(user_id)
        if user['is_blocked']:
            await update.message.reply_text("❌ Ваш аккаунт заблокирован. Обратитесь к администратору.")
            return
        
        await update.message.reply_text(
            f"👋 Добро пожаловать, {user['full_name']}!",
            reply_markup=get_main_menu_keyboard()
        )
    else:
        await update.message.reply_text(
            "👋 Здравствуйте! Вы не зарегистрированы. Введите Ваши полные ФИО для регистрации:",
            reply_markup=get_back_keyboard()
        )
        return WAITING_FOR_FULLNAME

async def handle_fullname(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🔙 Назад":
        await update.message.reply_text(
            "Регистрация отменена. Используйте /start для начала работы."
        )
        return ConversationHandler.END
    
    full_name = update.message.text
    user_id = update.effective_user.id
    
    role = user_db.add_user(user_id, full_name)
    
    # Логируем регистрацию
    action_logger.log_action(user_id, full_name, "Регистрация в системе")
    
    if role == 'admin':
        await update.message.reply_text(
            "🎉 Вы зарегистрированы как Администратор!",
            reply_markup=get_main_menu_keyboard()
        )
    else:
        await update.message.reply_text(
            "✅ Вы успешно зарегистрированы!",
            reply_markup=get_main_menu_keyboard()
        )
    
    return ConversationHandler.END

async def handle_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    
    if text == "📄 Соглашение":
        return await start_agreement_registration(update, context)
    elif text == "📨 Адвокатский запрос":
        return await start_query_registration(update, context)
    elif text == "🔙 Назад":
        await update.message.reply_text(
            "Главное меню:",
            reply_markup=get_main_menu_keyboard()
        )

# ===== РЕГИСТРАЦИЯ СОГЛАШЕНИЙ =====
async def start_agreement_registration(update: Update, context: ContextTypes.DEFAULT_TYPE):
    reg_number = generate_reg_number('agreement')
    context.user_data['current_agreement'] = {'reg_number': reg_number}
    
    user = user_db.get_user(update.effective_user.id)
    action_logger.log_action(update.effective_user.id, user['full_name'], f"Начало регистрации соглашения № {reg_number}")
    
    await update.message.reply_text(
        f"📄 Регистрация соглашения\n📋 Рег. №: {reg_number}\n\n"
        "Введите дату заключения соглашения (ДД.ММ.ГГГГ):",
        reply_markup=get_today_button_keyboard()
    )
    return WAITING_FOR_AGREEMENT_DATE

async def handle_agreement_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🔙 Назад":
        await update.message.reply_text(
            "❌ Регистрация соглашения отменена.",
            reply_markup=get_main_menu_keyboard()
        )
        context.user_data.pop('current_agreement', None)
        return ConversationHandler.END
    
    if update.message.text == "📅 Сегодня":
        date = datetime.now().strftime("%d.%m.%Y")
    else:
        date = update.message.text
    
    context.user_data['current_agreement']['conclusion_date'] = date
    # Дата регистрации всегда текущая дата
    context.user_data['current_agreement']['reg_date'] = datetime.now().strftime("%d.%m.%Y")
    
    await update.message.reply_text(
        "Сведения о доверителе (ФИО, наименование юр. лица, адрес):",
        reply_markup=get_back_keyboard()
    )
    return WAITING_FOR_CLIENT_INFO

async def handle_client_info(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🔙 Назад":
        await update.message.reply_text(
            "Введите дату заключения соглашения (ДД.ММ.ГГГГ):",
            reply_markup=get_today_button_keyboard()
        )
        return WAITING_FOR_AGREEMENT_DATE
    
    context.user_data['current_agreement']['client_info'] = update.message.text
    
    await update.message.reply_text(
        "Введите характер поручения (уголовное, гражданское, иное):",
        reply_markup=get_back_keyboard()
    )
    return WAITING_FOR_CASE_TYPE

async def handle_case_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🔙 Назад":
        await update.message.reply_text(
            "Сведения о доверителе (ФИО, наименование юр. лица, адрес):",
            reply_markup=get_back_keyboard()
        )
        return WAITING_FOR_CLIENT_INFO
    
    case_type = update.message.text
    context.user_data['current_agreement']['case_type'] = case_type
    
    # Сохраняем соглашение
    agreement_data = context.user_data['current_agreement']
    user = user_db.get_user(update.effective_user.id)
    
    record_data = [
        agreement_data['reg_number'],
        agreement_data['conclusion_date'],
        agreement_data['reg_date'],  # Автоматически установленная дата регистрации
        user['full_name'],
        agreement_data['client_info'],
        agreement_data['case_type']  # Характер поручения
    ]
    
    success = excel_manager.add_record('Журнал регистрации Соглашений', record_data)
    
    # Логируем действие
    if success:
        action_logger.log_action(
            update.effective_user.id, 
            user['full_name'], 
            f"Успешная регистрация соглашения № {agreement_data['reg_number']} (Характер: {case_type})"
        )
        
        # Выводим полную информацию о регистрации
        registration_info = (
            f"✅ Соглашение успешно зарегистрировано!\n\n"
            f"📋 Полная информация:\n"
            f"• Рег. №: {agreement_data['reg_number']}\n"
            f"• Дата заключения: {agreement_data['conclusion_date']}\n"
            f"• Дата регистрации: {agreement_data['reg_date']}\n"
            f"• Адвокат: {user['full_name']}\n"
            f"• Сведения о доверителе: {agreement_data['client_info']}\n"
            f"• Характер поручения: {case_type}"
        )
        
        await update.message.reply_text(
            registration_info,
            reply_markup=get_main_menu_keyboard()
        )
    else:
        action_logger.log_action(
            update.effective_user.id, 
            user['full_name'], 
            f"Ошибка при регистрации соглашения № {agreement_data['reg_number']}"
        )
        await update.message.reply_text(
            "❌ Ошибка при сохранении соглашения. Попробуйте еще раз.",
            reply_markup=get_main_menu_keyboard()
        )
    
    context.user_data.pop('current_agreement', None)
    return ConversationHandler.END

# ===== РЕГИСТРАЦИЯ АДВОКАТСКИХ ЗАПРОСОВ =====
async def start_query_registration(update: Update, context: ContextTypes.DEFAULT_TYPE):
    reg_number = generate_reg_number('query')
    context.user_data['current_query'] = {'reg_number': reg_number}
    
    user = user_db.get_user(update.effective_user.id)
    action_logger.log_action(update.effective_user.id, user['full_name'], f"Начало регистрации запроса № {reg_number}")
    
    await update.message.reply_text(
        f"📨 Регистрация адвокатского запроса\n📋 Рег. №: {reg_number}\n\n"
        "Введите дату запроса (ДД.ММ.ГГГГ):",
        reply_markup=get_today_button_keyboard()
    )
    return WAITING_FOR_QUERY_DATE

async def handle_query_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🔙 Назад":
        await update.message.reply_text(
            "❌ Регистрация запроса отменена.",
            reply_markup=get_main_menu_keyboard()
        )
        context.user_data.pop('current_query', None)
        return ConversationHandler.END
    
    if update.message.text == "📅 Сегодня":
        date = datetime.now().strftime("%d.%m.%Y")
    else:
        date = update.message.text
    
    context.user_data['current_query']['reg_date'] = date
    
    await update.message.reply_text(
        "Введите наименование адресата (органа гос. власти, иное):",
        reply_markup=get_back_keyboard()
    )
    return WAITING_FOR_RECIPIENT

async def handle_recipient(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🔙 Назад":
        await update.message.reply_text(
            "Введите дату запроса (ДД.ММ.ГГГГ):",
            reply_markup=get_today_button_keyboard()
        )
        return WAITING_FOR_QUERY_DATE
    
    context.user_data['current_query']['recipient'] = update.message.text
    
    # Сохраняем запрос
    query_data = context.user_data['current_query']
    user = user_db.get_user(update.effective_user.id)
    
    record_data = [
        query_data['reg_number'],
        query_data['reg_date'],
        query_data['recipient'],
        user['full_name']
    ]
    
    success = excel_manager.add_record('Журнал регистрации Адвокатских запросов', record_data)
    
    # Логируем действие
    if success:
        action_logger.log_action(
            update.effective_user.id, 
            user['full_name'], 
            f"Успешная регистрация запроса № {query_data['reg_number']}"
        )
        
        # Выводим полную информацию о регистрации
        registration_info = (
            f"✅ Адвокатский запрос успешно зарегистрирован!\n\n"
            f"📋 Полная информация:\n"
            f"• Рег. №: {query_data['reg_number']}\n"
            f"• Дата запроса: {query_data['reg_date']}\n"
            f"• Адресат: {query_data['recipient']}\n"
            f"• Адвокат: {user['full_name']}"
        )
        
        await update.message.reply_text(
            registration_info,
            reply_markup=get_main_menu_keyboard()
        )
    else:
        action_logger.log_action(
            update.effective_user.id, 
            user['full_name'], 
            f"Ошибка при регистрации запроса № {query_data['reg_number']}"
        )
        await update.message.reply_text(
            "❌ Ошибка при сохранении запроса. Попробуйте еще раз.",
            reply_markup=get_main_menu_keyboard()
        )
    
    context.user_data.pop('current_query', None)
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Операция отменена.",
        reply_markup=get_main_menu_keyboard()
    )
    return ConversationHandler.END

# ===== ОСНОВНАЯ ФУНКЦИЯ =====
def main():
    # Создаем приложение
    application = Application.builder().token(TELEGRAM_BOT_TOKEN).build()
    
    # Обработчик регистрации
    registration_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            WAITING_FOR_FULLNAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_fullname)],
        },
        fallbacks=[CommandHandler("cancel", cancel)]
    )
    
    # Обработчик соглашений
    agreement_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📄 Соглашение$"), start_agreement_registration)],
        states={
            WAITING_FOR_AGREEMENT_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_agreement_date)],
            WAITING_FOR_CLIENT_INFO: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_client_info)],
            WAITING_FOR_CASE_TYPE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_case_type)],
        },
        fallbacks=[CommandHandler("cancel", cancel)]
    )
    
    # Обработчик запросов
    query_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📨 Адвокатский запрос$"), start_query_registration)],
        states={
            WAITING_FOR_QUERY_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_query_date)],
            WAITING_FOR_RECIPIENT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_recipient)],
        },
        fallbacks=[CommandHandler("cancel", cancel)]
    )
    
    # Добавляем обработчики
    application.add_handler(registration_handler)
    application.add_handler(agreement_handler)
    application.add_handler(query_handler)
    
    # Основной обработчик текстовых сообщений
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_main_menu))
    
    # Запускаем бота
    logger.info("Бот запущен!")
    application.run_polling()

if __name__ == "__main__":
    main()